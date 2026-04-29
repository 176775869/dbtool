# coding=utf-8
"""
角色识别器 v2.0（Excel全局缓存，避免重复加载，修复先锋识别）
"""
import os
import openpyxl
from config_loader import load_config

# 全局缓存：整个进程生命周期只加载一次Excel
_REASON_CACHE = {}
_EXCEL_PATH = None

# 方向关键词（用于匹配涨停原因）
DIRECTION_KW = {
    '新能源/锂电': ['锂', '稀土', '固态电池', '刀片电池', '钠电池', '麒麟电池', '电池', '钨', '钴', '镍',
                    '盐湖', '能源金属', '磷化工'],
    '半导体/芯片产业链': ['半导体', '芯片', 'PCB', '封测', '集成电路', '光刻', 'MicroLED'],
    '算力/CPO': ['算力', 'CPO', '光模块', 'AI', '服务器', '数据中心'],
}

def load_reason_map(date_str):
    """从Excel加载指定日期的涨停原因（全局缓存）"""
    global _REASON_CACHE, _EXCEL_PATH

    if date_str in _REASON_CACHE:
        return _REASON_CACHE[date_str]

    if _EXCEL_PATH is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        _EXCEL_PATH = os.path.join(script_dir, '..', '..', 'backup', '交易复盘记录2026.xlsx')

    if not os.path.exists(_EXCEL_PATH):
        _REASON_CACHE[date_str] = {}
        return {}

    try:
        wb = openpyxl.load_workbook(_EXCEL_PATH, data_only=True)
        sheet_name = None
        for name in [date_str, date_str[2:] if len(date_str) == 8 else date_str]:
            if name in wb.sheetnames:
                sheet_name = name
                break

        if not sheet_name:
            wb.close()
            _REASON_CACHE[date_str] = {}
            return {}

        ws = wb[sheet_name]
        headers = [str(cell.value) if cell.value else '' for cell in ws[1]]
        code_col = reason_col = None
        for i, h in enumerate(headers):
            if h.strip() == '代码':
                code_col = i
            if '涨停原因类别' in h:
                reason_col = i

        if code_col is None or reason_col is None:
            wb.close()
            _REASON_CACHE[date_str] = {}
            return {}

        reason_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(code_col, reason_col):
                continue
            code = str(row[code_col]).strip().replace('SH', '').replace('SZ', '').replace('BJ', '')
            reason = str(row[reason_col]).strip() if row[reason_col] else ''
            if code and reason:
                reason_map[code] = reason

        wb.close()
        _REASON_CACHE[date_str] = reason_map
        return reason_map
    except Exception as e:
        _REASON_CACHE[date_str] = {}
        return {}


def identify_roles(direction_name, data):
    """识别主线方向的中军、连板先锋、弹性先锋、趋势先锋"""
    limit_stocks = data.get('limit_stocks', [])
    top20 = data.get('top20', [])
    mid_caps = data.get('mid_caps', [])
    qs = data.get('qs_stocks', [])
    date_str = data.get('date', '')

    reason_map = load_reason_map(date_str) if date_str else {}
    kw_list = DIRECTION_KW.get(direction_name, [direction_name])

    # ---- 1. 中军：从 top20 + mid_caps 中选市值 > 200亿 且 涨 > 3% ----
    all_mid = top20 + mid_caps
    seen = set()
    mids = []
    for obj in all_mid:
        code = obj.get('code', '')
        if code in seen:
            continue
        if obj.get('market_cap', 0) >= 200 and obj.get('pct', 0) >= 3:
            mids.append(obj)
            seen.add(code)

    # 兜底：放宽条件
    if not mids:
        for obj in all_mid:
            code = obj.get('code', '')
            if code in seen:
                continue
            if obj.get('market_cap', 0) >= 100 and obj.get('pct', 0) >= 1:
                mids.append(obj)
                seen.add(code)

    mids.sort(key=lambda x: x.get('market_cap', 0), reverse=True)
    mids = mids[:3]

    # ---- 2. 筛选涨停股（优先涨停原因表，兜底方向关键词） ----
    rel_limit = []
    for st in limit_stocks:
        code = st.get('code', '')
        reason = reason_map.get(code, '')
        if reason and any(kw in reason for kw in kw_list):
            rel_limit.append(st)

    # 兜底：用概念板块匹配
    if not rel_limit:
        for st in limit_stocks:
            sector = st.get('sector', '')
            if any(kw in sector for kw in kw_list):
                rel_limit.append(st)

    # ---- 3. 连板先锋（10cm）+ 弹性先锋（20cm） ----
    def score_lb(st):
        t = st.get('turnover', 0)
        ts = 10 if 3 <= t <= 7 else (7 if 7 < t <= 15 else (3 if t < 3 else 2))
        try:
            h, m = map(int, st.get('first_time', '15:00').split(':'))
            mins = h * 60 + m
        except:
            mins = 999
        ti = (10 if mins <= 570 else 9 if mins <= 600 else 7 if mins <= 630 else 5 if mins <= 690 else 2)
        sr = st.get('seal_amount', 0) / st.get('amount', 1) if st.get('amount', 0) > 0 else 0
        se = 9 if sr > 3 else (7 if sr > 2 else (5 if sr > 1 else 3))
        is_10cm = not st['code'].startswith(('300', '688', '301'))
        return ts * 0.3 + ti * 0.3 + se * 0.2 + (5 if is_10cm else -5) * 0.2, st

    lianban, elastic = None, None
    if rel_limit:
        scored = [score_lb(s) for s in rel_limit]
        scored.sort(key=lambda x: x[0], reverse=True)
        top1 = scored[0][1]
        if top1['code'].startswith(('300', '688', '301')):
            elastic = top1
            for sc, st in scored:
                if not st['code'].startswith(('300', '688', '301')):
                    lianban = st
                    break
        else:
            lianban = top1
            for sc, st in scored:
                if st['code'].startswith(('300', '688', '301')):
                    elastic = st
                    break

    # ---- 4. 趋势先锋 ----
    trend = None
    rel_qs = []
    for q in qs:
        code = q.get('code', '')
        r = reason_map.get(code, '')
        if r and any(kw in r for kw in kw_list):
            rel_qs.append(q)
    if not rel_qs:
        rel_qs = qs

    cands = [q for q in rel_qs if q.get('lb', 0) >= 1.5]
    valid = [q for q in cands if not any(ls['code'] == q['code'] and ls['lianban'] > 2 for ls in limit_stocks)]
    if valid:
        valid.sort(key=lambda x: x.get('trend_score', 0), reverse=True)
        trend = valid[0]

    return {
        'mid_cap': mids,
        'lianban_pioneer': lianban,
        'elastic_pioneer': elastic,
        'trend_pioneer': trend
    }